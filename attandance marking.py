def sorted():
     import xlrd
     import openpyxl 
     import datetime
     loc = (r'C:\Users\ADMIN\Documents\FaceRecognition-master\dimension.xlsx')  
     wb = xlrd.open_workbook(loc) 
     sheet = wb.sheet_by_index(0) 
     sheet.cell_value(0, 0)
     l=[]
     for i in range(sheet.nrows):
          l.append(sheet.cell_value(i, 0))
     print(l)
     temp=[]
     for i in range(1,len(l)):
          if l[i] not in temp:
               temp.append(l[i])
     a=0
     f=[]
     for j in temp:
          for i in range(1,len(l)):    
               if(l[i]==j):
                    a=a+1    
          if(a>5):
               f.append(j)
               a=0
          if(a<5):
               a=0
     print(f)
     import openpyxl 
     wbkName = r'C:\Users\ADMIN\Desktop\dimension.xlsx'
     wbk = openpyxl.load_workbook(wbkName)
     sheet = wbk.worksheets[0]
     p=sheet.max_row
     print(p)
     print(len(f))
     s=p+len(f)
     print(s)
     for wks in wbk.worksheets:
             for i in range(p+1,s+1):
                     for j in range(1,5):
                    
                             wks.cell(row=i,column=1).value=f[i+1-s]
                             now= datetime.datetime.now()
                             wks.cell(row=i,column=2).value=now.hour
                             wks.cell(row=i,column=3).value=now.minute
                             wks.cell(row=i,column=4).value=now.second
                             wks.cell(row=i,column=5).value=datetime.date.today()
			
			
     wbk.save(wbkName)
     wbk.close
