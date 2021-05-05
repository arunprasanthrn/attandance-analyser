
def time():
    
    import openpyxl
    import datetime
    i=2



 
    wb = openpyxl.Workbook() 
 
    sheet = wb.active

    sheet.cell(row = 1, column = 1).value = 'Name'
    sheet.cell(row = 1, column = 2).value = 'Hour'
    sheet.cell(row =1, column = 3).value ='Minute'
    sheet.cell(row = 1, column = 4).value = 'Second'
    sheet.cell(row = 1, column = 5).value = 'Date'




    

# set the height of the row 
    sheet.row_dimensions[1].height = 20

# set the width of the column 
    sheet.column_dimensions['B'].width = 20

# save the file 
    wb.save('dimension.xlsx') 

time()
