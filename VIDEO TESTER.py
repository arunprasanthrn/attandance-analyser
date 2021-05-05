import os
import cv2
import numpy as np
import faceRecognition as fr
import openpyxl
import datetime


#This module captures images via webcam and performs face recognition
face_recognizer = cv2.face.LBPHFaceRecognizer_create()
face_recognizer.read('trainingData.yml')#Load saved training data
wb = openpyxl.Workbook()
sheet = wb.active
name = {0 : "kishore",1 : "arun",2:"jegadheesh",3:"kp",4:"nila"}


cap=cv2.VideoCapture(0)
i=2


for j in range(0,1000):
    ret,test_img=cap.read()# captures frame and returns boolean value and captured image
    faces_detected,gray_img=fr.faceDetection(test_img)



    for (x,y,w,h) in faces_detected:
      cv2.rectangle(test_img,(x,y),(x+w,y+h),(255,0,0),thickness=7)

    resized_img = cv2.resize(test_img, (1000, 700))
    cv2.imshow('face detection Tutorial ',resized_img)
    cv2.waitKey(10)


    for face in faces_detected:
        (x,y,w,h)=face
        roi_gray=gray_img[y:y+w, x:x+h]
        label,confidence=face_recognizer.predict(roi_gray)#predicting the label of given image
        print("confidence:",confidence)
        print("label:",label)
        fr.draw_rect(test_img,face)
        predicted_name=name[label]
        if confidence < 37:#If confidence less than 37 then don't print predicted face text on screen
           fr.put_text(test_img,predicted_name,x,y)
           now= datetime.datetime.now()
           hour=now.hour
           minute=now.minute
           second=now.second
           sheet.cell(row = i, column = 1).value =predicted_name
           sheet.cell(row = i, column = 2).value = hour
           sheet.cell(row = i, column = 3).value =minute
           sheet.cell(row = i, column = 4).value = second
           i=i+1
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions['B'].width = 20
        wb.save('dimension.xlsx') 


    resized_img = cv2.resize(test_img, (1000, 700))
    cv2.imshow('face recognition tutorial ',resized_img)
    cv2.waitKey(10)#wait until 'q' key is pressed
       
    



cap.release()
cv2.destroyAllWindows

