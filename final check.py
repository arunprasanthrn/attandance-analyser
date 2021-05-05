#To create a empty document
def doc():
    
    import openpyxl
    import datetime
    i=2
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.cell(row = 1, column = 1).value = 'Name'
    sheet.cell(row = 1, column = 2).value = 'Hour'
    sheet.cell(row =1, column = 3).value ='Minute'
    sheet.cell(row = 1, column = 4).value = 'Second'
    sheet.cell(row = 1, column = 5).value = 'Date'
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions['B'].width = 20
    wb.save('Present.xlsx')
    
    import openpyxl
    import datetime
    i=2
    wb = openpyxl.Workbook() 
    sheet = wb.active
    sheet.cell(row = 1, column = 1).value = 'Name'
    sheet.cell(row = 1, column = 2).value = 'Hour'
    sheet.cell(row =1, column = 3).value ='Minute'
    sheet.cell(row = 1, column = 4).value = 'Second'
    sheet.cell(row = 1, column = 5).value = 'Date'
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions['B'].width = 20
    wb.save('Abscent.xlsx')




    

#To take attandance

def present():
  
    import xlrd
    import openpyxl 
    import datetime
    loc = (r'C:\Users\ADMIN\Documents\FaceRecognition-master\dimension.xlsx')  
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0) 
    sheet.cell_value(0, 0)
    l=[]
    for i in range(sheet.nrows):
         l.append(sheet.cell_value(i, 0))
    print(l)
    temp=[]
    for i in range(1,len(l)):
         if l[i] not in temp:
              temp.append(l[i])
    a=0
    f=[]
    for j in temp:
         for i in range(1,len(l)):    
              if(l[i]==j):
                   a=a+1    
         if(a>5):
              f.append(j)
              a=0
         if(a<5):
              a=0
    print(f)
    a=[]
    total=['kishore','arun','jegadheesh','prakash']
    for i in total:
        if i  not in f:
            a.append(i)
    print(a)    

    import openpyxl 
    wbkName = r'C:\Users\ADMIN\Desktop\present.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    sheet = wbk.worksheets[0]
    p=sheet.max_row
    print(p)
    print(len(f))
    s=p+len(f)
    print(s)
    for wks in wbk.worksheets:
              for i in range(p+1,s+1):
                        for j in range(1,5):
                        
                             wks.cell(row=i,column=1).value=f[i+1-s]
                             now= datetime.datetime.now()
                             wks.cell(row=i,column=2).value=now.hour
                             wks.cell(row=i,column=3).value=now.minute
                             wks.cell(row=i,column=4).value=now.second
                             wks.cell(row=i,column=5).value=datetime.date.today()
                            
                            
    wbk.save(wbkName)
    wbk.close
def abscent():
    import xlrd
    import openpyxl 
    import datetime
    loc = (r'C:\Users\ADMIN\Documents\FaceRecognition-master\dimension.xlsx')  
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0) 
    sheet.cell_value(0, 0)
    l=[]
    for i in range(sheet.nrows):
         l.append(sheet.cell_value(i, 0))
    print(l)
    temp=[]
    for i in range(1,len(l)):
         if l[i] not in temp:
              temp.append(l[i])
    a=0
    f=[]
    for j in temp:
         for i in range(1,len(l)):    
              if(l[i]==j):
                   a=a+1    
         if(a>5):
              f.append(j)
              a=0
         if(a<5):
              a=0
    print(f)
    a=[]
    total=['kishore','arun','jegadheesh','prakash']
    for i in total:
        if i  not in f:
            a.append(i)
    print(a)  
    
    import openpyxl 
    wbkName = r'C:\Users\ADMIN\Desktop\Abscent.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    sheet = wbk.worksheets[0]
    q=sheet.max_row
    print(q)
    print(len(a))
    m=q+len(a)
    print(m)
    for wks in wbk.worksheets:
              for i in range(q+1,m+1):
                        for j in range(1,5):
                        
                             wks.cell(row=i,column=1).value=a[i-m]
                             now= datetime.datetime.now()
                             wks.cell(row=i,column=2).value=now.hour
                             wks.cell(row=i,column=3).value=now.minute
                             wks.cell(row=i,column=4).value=now.second
                             wks.cell(row=i,column=5).value=datetime.date.today()
                            
                            
    wbk.save(wbkName)
    wbk.close
    
    
def start():
    def hi():
        import datetime
        
        now= datetime.datetime.now()
        hour=now.hour
        minute=now.minute
        second=now.second
        if((hour==9 and minute==00) or (hour==9 and minute==5) or (hour==11 and minute==10)):
            import os
            import cv2
            import numpy as np
            import faceRecognition as fr
            import openpyxl
            import datetime
            #This module captures images via webcam and performs face recognition
            face_recognizer = cv2.face.LBPHFaceRecognizer_create()
            face_recognizer.read('trainingData.yml')#Load saved training data
            wb = openpyxl.Workbook()
            sheet = wb.active
            name = {0 : "kishore",1 : "arun",2:"jegadheesh",3:"nila"}


            cap=cv2.VideoCapture(0)
            i=2
            
            for j in range(0,1000):
                ret,test_img=cap.read()# captures frame and returns boolean value and captured image
                faces_detected,gray_img=fr.faceDetection(test_img)



                for (x,y,w,h) in faces_detected:
                    
                    cv2.rectangle(test_img,(x,y),(x+w,y+h),(255,0,0),thickness=7)

                    resized_img = cv2.resize(test_img, (1000, 700))
                    cv2.imshow('face detection Tutorial ',resized_img)
                    cv2.waitKey(10)


                for face in faces_detected:
                    (x,y,w,h)=face
                    roi_gray=gray_img[y:y+w, x:x+h]
                    label,confidence=face_recognizer.predict(roi_gray)#predicting the label of given image
                    print("confidence:",confidence)
                    print("label:",label)
                    fr.draw_rect(test_img,face)
                    predicted_name=name[label]
                    if confidence < 90:
                        #If confidence less than 37 then don't print predicted face text on screen
                        fr.put_text(test_img,predicted_name,x,y)
                        now= datetime.datetime.now()
                        hour=now.hour
                        minute=now.minute
                        second=now.second
                        sheet.cell(row = i, column = 1).value =predicted_name
                        sheet.cell(row = i, column = 2).value = hour
                        sheet.cell(row = i, column = 3).value =minute
                        sheet.cell(row = i, column = 4).value = second
                        i=i+1
                    sheet.row_dimensions[1].height = 20
                    sheet.column_dimensions['B'].width = 20
                    wb.save('dimension.xlsx') 


                resized_img = cv2.resize(test_img, (1000, 700))
                cv2.imshow('face recognition tutorial ',resized_img)
                cv2.waitKey(10)#wait until 'q' key is pressed
           
        



            cap.release()
            cv2.destroyAllWindows
        else:
            print("no")


       
    while True:
        hi()
        










import tkinter as tk
root = tk.Tk()

button1 = tk.Button(root, text='start', width=25, command=root.destroy)
button1.pack()
button1.config(command=start)
button2 = tk.Button(root, text='Present', width=25, command=root.destroy)
button2.pack()
button2.config(command=present)
button2 = tk.Button(root, text='Abscent', width=25, command=root.destroy)
button2.pack()
button2.config(command=abscent)
button2 = tk.Button(root, text='Reset', width=25, command=root.destroy)
button2.pack()
button2.config(command=doc)
root.mainloop()
