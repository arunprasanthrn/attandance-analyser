import xlrd 
  
loc = (r'C:\Users\ADMIN\Documents\FaceRecognition-master\dimension.xlsx') 
  
wb = xlrd.open_workbook(loc) 
sheet = wb.sheet_by_index(0) 
sheet.cell_value(0, 0)
l=[]
for i in range(sheet.nrows):
     l.append(sheet.cell_value(i, 0))
print(l)
temp=[]
for i in range(1,len(l)):
     if l[i] not in temp:
          temp.append(l[i])

a=0
f=[]
for j in temp:
     for i in range(1,len(l)):
     
          if(l[i]==j):
               a=a+1
     
     if(a>5):
          f.append(j)
          a=0
     if(a<5):
          a=0
print(f)


#TO SHOTLIST


import openpyxl
import datetime
i=2



 
wb = openpyxl.Workbook() 
 
sheet = wb.active 

for i in range(0,len(f)):
    now= datetime.datetime.now()
    hour=now.hour
    minute=now.minute
    second=now.second
    sheet.cell(row = i+1, column = 1).value = f[i]
    sheet.cell(row = i+1, column = 2).value = hour
    sheet.cell(row = i+1, column = 3).value =minute
    sheet.cell(row = i+1, column = 4).value = second
    i=i+1
    if(i>5):
        break
    

# set the height of the row 
sheet.row_dimensions[1].height = 20

# set the width of the column 
sheet.column_dimensions['B'].width = 20

# save the file 
wb.save('dimension.xlsx') 

          
               

               

    

        
    
    
